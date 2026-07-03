"""
cimt_build.py
=============
Reads every CIMT zip file in your CIMTD folder, builds a full 2012-2026
pivot table per commodity (Exports and Imports), and saves one Excel
workbook per commodity.

Run after cimt_download.py has populated the CIMTD folder.

OUTPUT (per commodity, e.g. CIMT_Canola_20260411.xlsx)
-------------------------------------------------------
  Exports    Years x Months pivot  (KMT, whole numbers)
  Imports    Years x Months pivot  (KMT, whole numbers)
  Raw Data   Flat table with Country, HS6, Year, Month, KMT, Value(CAD)
             — use column filters to slice by country / HS6 code
"""

import zipfile
import calendar
from pathlib import Path
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

CIMTD = Path("data")  # <-- set this to your own folder"")

COMMODITY_CONFIG = {
    "Amber Durum": {
        "hs6_codes": [100111, 100119],
        "crop_year_start_month": 8,
        "top_destinations": 10,
    },
    "Wheat (Ex-Durum)": {
        "hs6_codes": [100191, 100199],
        "crop_year_start_month": 8,
        "top_destinations": 20,
    },
    "Canola": {
        "hs6_codes": [120510, 120590],
        "crop_year_start_month": 8,
        "top_destinations": 10,
    },
    "Soybeans": {
        "hs6_codes": [120110, 120190],
        "crop_year_start_month": 9,
        "top_destinations": 10,
    },
    "Corn": {
        "hs6_codes": [100510, 100590],
        "crop_year_start_month": 8,
        "top_destinations": 10,
    },
    "Barley": {
        "hs6_codes": [100310, 100390],
        "crop_year_start_month": 8,
        "top_destinations": 10,
    },
    "Red Lentils": {
        "hs6_codes": [71340, 71341],
        "crop_year_start_month": 8,
        "top_destinations": 10,
    },
    "Yellow Peas": {
        "hs6_codes": [71310, 71390],
        "crop_year_start_month": 8,
        "top_destinations": 10,
    },
}

# =============================================================================
# REGIONAL MAPPING
# =============================================================================
# Countries mapped to trade blocs / geographic regions.
# Any country not listed falls into "Other".

REGION_MAP = {
    # ── European Union (EU27)
    "Austria":              "Europe (+UK & Non-EU)",
    "Belgium":              "Europe (+UK & Non-EU)",
    "Bulgaria":             "Europe (+UK & Non-EU)",
    "Croatia":              "Europe (+UK & Non-EU)",
    "Cyprus":               "Europe (+UK & Non-EU)",
    "Czechia":              "Europe (+UK & Non-EU)",
    "Denmark":              "Europe (+UK & Non-EU)",
    "Estonia":              "Europe (+UK & Non-EU)",
    "Finland":              "Europe (+UK & Non-EU)",
    "France":               "Europe (+UK & Non-EU)",
    "Germany":              "Europe (+UK & Non-EU)",
    "Greece":               "Europe (+UK & Non-EU)",
    "Hungary":              "Europe (+UK & Non-EU)",
    "Ireland":              "Europe (+UK & Non-EU)",
    "Italy":                "Europe (+UK & Non-EU)",
    "Latvia":               "Europe (+UK & Non-EU)",
    "Lithuania":            "Europe (+UK & Non-EU)",
    "Luxembourg":           "Europe (+UK & Non-EU)",
    "Malta":                "Europe (+UK & Non-EU)",
    "Netherlands":          "Europe (+UK & Non-EU)",
    "Poland":               "Europe (+UK & Non-EU)",
    "Portugal":             "Europe (+UK & Non-EU)",
    "Romania":              "Europe (+UK & Non-EU)",
    "Slovakia":             "Europe (+UK & Non-EU)",
    "Slovenia":             "Europe (+UK & Non-EU)",
    "Spain":                "Europe (+UK & Non-EU)",
    "Sweden":               "Europe (+UK & Non-EU)",
    # ── Non-EU Europe
    "United Kingdom":       "Europe (+UK & Non-EU)",               # included with EU for trade analysis purposes
    "Norway":               "Europe (+UK & Non-EU)",
    "Switzerland":          "Europe (+UK & Non-EU)",
    "Ukraine":              "Europe (+UK & Non-EU)",
    "Turkey":               "Europe (+UK & Non-EU)",
    "Serbia":               "Europe (+UK & Non-EU)",
    "Albania":              "Europe (+UK & Non-EU)",
    "Bosnia and Herzegovina": "Europe (+UK & Non-EU)",
    "Macedonia, North":     "Europe (+UK & Non-EU)",
    "Moldova, Republic of": "Europe (+UK & Non-EU)",
    "Belarus":              "Europe (+UK & Non-EU)",
    "Georgia":              "Europe (+UK & Non-EU)",
    "Armenia":              "Europe (+UK & Non-EU)",
    "Azerbaijan":           "Europe (+UK & Non-EU)",
    "Iceland":              "Europe (+UK & Non-EU)",
    "Russian Federation":   "Europe (+UK & Non-EU)",
    "Montenegro":           "Europe (+UK & Non-EU)",
    "Kosovo":               "Europe (+UK & Non-EU)",
    # ── North America
    "United States of America": "North America",
    "Mexico":               "North America",
    # ── LATAM (Caribbean + Central + South America)
    "Cuba":                 "LATAM",
    "Haiti":                "LATAM",
    "Dominican Republic":   "LATAM",
    "Jamaica":              "LATAM",
    "Trinidad and Tobago":  "LATAM",
    "Barbados":             "LATAM",
    "Bahamas":              "LATAM",
    "Guyana":               "LATAM",
    "Suriname":             "LATAM",
    "Guatemala":            "LATAM",
    "Honduras":             "LATAM",
    "El Salvador":          "LATAM",
    "Nicaragua":            "LATAM",
    "Costa Rica":           "LATAM",
    "Panama":               "LATAM",
    "Colombia":             "LATAM",
    "Venezuela":            "LATAM",
    "Ecuador":              "LATAM",
    "Peru":                 "LATAM",
    "Bolivia":              "LATAM",
    "Brazil":               "LATAM",
    "Paraguay":             "LATAM",
    "Uruguay":              "LATAM",
    "Argentina":            "LATAM",
    "Chile":                "LATAM",
    "Belize":               "LATAM",
    "Antigua and Barbuda":  "LATAM",
    "Dominica":             "LATAM",
    "Grenada":              "LATAM",
    "Saint Kitts and Nevis": "LATAM",
    "Saint Lucia":          "LATAM",
    "Saint Vincent and the Grenadines": "LATAM",
    # ── Middle East & North Africa (MENA)
    "Algeria":              "MENA",
    "Morocco":              "MENA",
    "Tunisia":              "MENA",
    "Libya":                "MENA",
    "Egypt":                "MENA",
    "Sudan":                "MENA",
    "Saudi Arabia":         "MENA",
    "United Arab Emirates": "MENA",
    "Kuwait":               "MENA",
    "Qatar":                "MENA",
    "Bahrain":              "MENA",
    "Oman":                 "MENA",
    "Yemen":                "MENA",
    "Jordan":               "MENA",
    "Lebanon":              "MENA",
    "Syria":                "MENA",
    "Iraq":                 "MENA",
    "Iran":                 "MENA",
    "Israel":               "MENA",
    "Palestine, State of":  "MENA",
    # ── Sub-Saharan Africa
    "Nigeria":              "Africa",
    "Ethiopia":             "Africa",
    "Kenya":                "Africa",
    "Tanzania, United Republic of": "Africa",
    "Ghana":                "Africa",
    "Cameroon":             "Africa",
    "Senegal":              "Africa",
    "Mozambique":           "Africa",
    "Zimbabwe":             "Africa",
    "Zambia":               "Africa",
    "Angola":               "Africa",
    "Uganda":               "Africa",
    "Côte d'Ivoire":        "Africa",
    "Madagascar":           "Africa",
    "Malawi":               "Africa",
    "Rwanda":               "Africa",
    "Benin":                "Africa",
    "Togo":                 "Africa",
    "Sierra Leone":         "Africa",
    "Liberia":              "Africa",
    "South Africa, Republic of": "Africa",
    "Namibia":              "Africa",
    "Botswana":             "Africa",
    "Mauritius":            "Africa",
    "Réunion":              "Africa",
    "Djibouti":             "Africa",
    "Somalia":              "Africa",
    "Congo, Republic of the": "Africa",
    "Congo, Democratic Republic of the": "Africa",
    "Gabon":                "Africa",
    "Guinea":               "Africa",
    "Mali":                 "Africa",
    "Niger":                "Africa",
    "Burkina Faso":         "Africa",
    "Chad":                 "Africa",
    "Eritrea":              "Africa",
    "South Sudan":          "Africa",
    # ── Asia (East + Southeast + South)
    "China":                "Asia",
    "Japan":                "Asia",
    "Korea, South":         "Asia",
    "Korea, North":         "Asia",
    "Taiwan":               "Asia",
    "Hong Kong":            "Asia",
    "Mongolia":             "Asia",
    "Viet Nam":             "Asia",
    "Thailand":             "Asia",
    "Indonesia":            "Asia",
    "Malaysia":             "Asia",
    "Philippines":          "Asia",
    "Singapore":            "Asia",
    "Myanmar":              "Asia",
    "Cambodia":             "Asia",
    "Laos":                 "Asia",
    "Bangladesh":           "Asia",
    "India":                "Asia",
    "Pakistan":             "Asia",
    "Sri Lanka":            "Asia",
    "Nepal":                "Asia",
    "Afghanistan":          "Asia",
    "Kazakhstan":           "Asia",
    "Uzbekistan":           "Asia",
    "Turkmenistan":         "Asia",
    "Kyrgyzstan":           "Asia",
    "Tajikistan":           "Asia",
    # ── Oceania
    "Australia":            "Oceania",
    "New Zealand":          "Oceania",
    "Papua New Guinea":     "Oceania",
    "Fiji":                 "Oceania",
}

REGION_ORDER = [
    "North America", "LATAM", "Europe (+UK & Non-EU)",
    "MENA", "Africa", "Asia", "Oceania", "Other"
]

MONTH_ABBR = [calendar.month_abbr[m] for m in range(1, 13)]

# =============================================================================
# DATA LOADING
# =============================================================================

def detect_trade_type(zip_path: Path) -> str:
    name = zip_path.name.upper()
    if "IMP" in name:
        return "Import"
    if "EXP" in name or "TOT" in name:
        return "Export"
    raise ValueError(f"Cannot detect trade type from: {zip_path.name}")


def parse_country_lookup(zf: zipfile.ZipFile) -> dict:
    """
    Parses ODPF_6_CtyDesc.TXT.
    Line format: CC NNN     DDDDDD DDDDDD English Name        French Name
    English name reliably starts at position 25.
    """
    import re as _re
    fname = next((n for n in zf.namelist() if "CtyDesc" in n), None)
    if not fname:
        return {}
    mapping = {}
    with zf.open(fname) as f:
        for raw in f:
            line = raw.decode("latin-1")
            if len(line) < 30:
                continue
            m = _re.match(r'(\w{2})\s+\d+\s+\d{6}\s+\d{6}\s+(.+?)(?:\s{2,}|\s*$)', line)
            if m:
                code    = m.group(1).strip()
                english = m.group(2).strip()
                if code and english:
                    mapping[code] = english
    return mapping


def find_hs6_csv(zf: zipfile.ZipFile, trade_type: str) -> str:
    """Locate ODPFN015 (imports) or ODPFN019 (exports) inside the zip."""
    prefix = "ODPFN015" if trade_type == "Import" else "ODPFN019"
    match  = next(
        (n for n in zf.namelist()
         if Path(n).name.startswith(prefix) and n.endswith(".csv")),
        None
    )
    if not match:
        # Fallback: skip the known non-HS6 files
        skip = {"ODPFN021", "ODPFN022", "ODPFN017", "ODPFN014"}
        match = next(
            (n for n in zf.namelist()
             if n.endswith(".csv") and not any(s in Path(n).name for s in skip)),
            None
        )
    if not match:
        raise FileNotFoundError(
            f"No HS6 CSV found in {zf.filename}. Contents: {zf.namelist()}"
        )
    return match


def read_zip(zip_path: Path, hs6_filter: set = None) -> pd.DataFrame:
    """
    Read one zip, return clean flat DataFrame.
    hs6_filter: if provided, only rows with matching HS6 codes are kept.
                Pass ALL commodity HS6 codes to dramatically reduce memory usage.
    """
    trade_type = detect_trade_type(zip_path)
    print(f"  {zip_path.name:<40} [{trade_type}]", end="  ", flush=True)

    with zipfile.ZipFile(zip_path) as zf:
        country_map = parse_country_lookup(zf)
        csv_path    = find_hs6_csv(zf, trade_type)

        if hs6_filter:
            # Read in chunks, filter early — avoids loading 4M rows into RAM
            chunks = []
            with zf.open(csv_path) as f:
                for chunk in pd.read_csv(f, encoding="latin-1",
                                         low_memory=False, chunksize=200_000):
                    # Find HS6 column name in this chunk
                    hs6_col = next((c for c in chunk.columns if c.strip() == "HS6"), None)
                    if hs6_col:
                        chunk[hs6_col] = pd.to_numeric(chunk[hs6_col], errors="coerce")
                        chunk = chunk[chunk[hs6_col].isin(hs6_filter)]
                    chunks.append(chunk)
            df = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
        else:
            with zf.open(csv_path) as f:
                df = pd.read_csv(f, encoding="latin-1", low_memory=False)

    print(f"{len(df):>10,} rows (filtered)")

    # ── Normalise bilingual column names
    rename = {}
    for col in df.columns:
        c = col.lower()
        if "yearmonth" in c or "annee" in c or "anné" in c:
            rename[col] = "YearMonth"
        elif col.strip() == "HS6":
            rename[col] = "HS6"
        elif "country" in c or "pays" in c:
            rename[col] = "CountryCode"
        elif "value" in c or "valeur" in c:
            rename[col] = "ValueCAD"
        elif "quantity" in c or "quantit" in c:
            rename[col] = "Quantity"
        elif "unit" in c:
            rename[col] = "UOM"
    df = df.rename(columns=rename)

    # ── Parse YearMonth (YYYYMM) -> Year, Month
    ym        = df["YearMonth"].astype(int).astype(str).str.zfill(6)
    df["Year"]  = ym.str[:4].astype(int)
    df["Month"] = ym.str[4:6].astype(int)

    # ── HS6 as integer
    df["HS6"] = pd.to_numeric(df["HS6"], errors="coerce").astype("Int64")

    # ── Country code -> full English name
    df["Country"] = df["CountryCode"].map(country_map).fillna(df["CountryCode"])

    # ── Quantity -> KMT  (per-row UOM conversion)
    # KGM/KGE = kilograms  -> divide by 1,000,000
    # TNE = tonnes         -> divide by 1,000
    # anything else (NMB, LTR, ...) = not a weight -> NaN (excluded)
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    if "UOM" in df.columns:
        uom = df["UOM"].astype(str).str.strip().str.upper()
        df["KMT"] = float("nan")
        df.loc[uom.isin(["KGM", "KGE"]), "KMT"] = df.loc[uom.isin(["KGM", "KGE"]), "Quantity"] / 1_000_000
        df.loc[uom == "TNE", "KMT"] = df.loc[uom == "TNE", "Quantity"] / 1_000
        df["KMT"] = pd.to_numeric(df["KMT"], errors="coerce")
    else:
        df["KMT"] = df["Quantity"] / 1_000  # fallback

    df["ValueCAD"]  = pd.to_numeric(df["ValueCAD"], errors="coerce").fillna(0)
    df["TradeType"] = trade_type

    # ── Aggregate to country level (imports have province x state rows per country)
    df = df.dropna(subset=["KMT"])
    df = (
        df.groupby(["Year", "Month", "TradeType", "HS6", "Country"], as_index=False)
        .agg({"KMT": "sum", "ValueCAD": "sum"})
    )
    return df[["Year", "Month", "TradeType", "HS6", "Country", "KMT", "ValueCAD"]]


def build_master_country_lookup(zips: list) -> dict:
    """
    Builds one definitive code->name lookup by scanning CtyDesc.TXT
    from every available zip. Later zips overwrite earlier ones,
    so the most recent naming convention wins.
    """
    import re as _re
    master = {}
    for z in zips:
        try:
            with zipfile.ZipFile(z) as zf:
                fname = next((n for n in zf.namelist() if "CtyDesc" in n), None)
                if not fname:
                    continue
                with zf.open(fname) as f:
                    for raw in f:
                        line = raw.decode("latin-1")
                        if len(line) < 30:
                            continue
                        m = _re.match(
                            r'(\w{2})\s+\d+\s+\d{6}\s+\d{6}\s+(.+?)(?:\s{2,}|\s*$)',
                            line
                        )
                        if m:
                            code    = m.group(1).strip()
                            english = m.group(2).strip()
                            if code and english:
                                master[code] = english
        except Exception:
            pass
    return master


def load_all_zips(folder: Path) -> pd.DataFrame:
    """Load every CIMT zip in the folder, concatenate, deduplicate."""
    zips = sorted(folder.glob("CIMT-CICM_*.zip"))
    if not zips:
        raise RuntimeError(f"No CIMT zip files found in {folder}")

    # Collect all HS6 codes across all commodities for early filtering
    all_hs6 = set()
    for cfg in COMMODITY_CONFIG.values():
        all_hs6.update(cfg["hs6_codes"])
    print(f"Filtering to {len(all_hs6)} HS6 codes across {len(COMMODITY_CONFIG)} commodities")
    print(f"Found {len(zips)} zip files in {folder}\n")

    # Build one master country lookup from all available zips
    master_lookup = build_master_country_lookup(zips)
    print(f"  Master country lookup: {len(master_lookup)} codes resolved\n")

    frames = []
    for z in zips:
        try:
            frames.append(read_zip(z, hs6_filter=all_hs6))
        except Exception as e:
            print(f"  ERROR {z.name}: {e}")

    if not frames:
        raise RuntimeError("No data loaded.")

    df = pd.concat(frames, ignore_index=True)

    # Final pass: resolve any remaining 2-letter codes using master lookup
    # (happens when per-zip lookup fails on older zip formats)
    unresolved = df["Country"].str.len() == 2
    if unresolved.any():
        n_before = unresolved.sum()
        df.loc[unresolved, "Country"] = (
            df.loc[unresolved, "Country"].map(master_lookup)
            .fillna(df.loc[unresolved, "Country"])
        )
        n_after = (df["Country"].str.len() == 2).sum()
        print(f"  Country resolution: {n_before} 2-letter codes -> "
              f"{n_after} remaining unresolved")

    df = df.drop_duplicates(
        subset=["Year", "Month", "TradeType", "HS6", "Country"]
    ).sort_values(["TradeType", "Year", "Month", "Country"])

    print(f"\n  Combined: {len(df):,} rows | "
          f"Years {df['Year'].min()}-{df['Year'].max()} | "
          f"{df['HS6'].nunique()} HS6 codes\n")
    return df


# =============================================================================
# PIVOT BUILDER
# =============================================================================

def build_pivot(df: pd.DataFrame, trade_type: str) -> pd.DataFrame:
    """
    Returns Years (rows) x Jan-Dec (cols) pivot of KMT.
    Values are rounded to whole integers, matching the target display format.
    Empty months are left blank (None).
    """
    sub = df[df["TradeType"] == trade_type]
    if sub.empty:
        return pd.DataFrame()

    piv = (
        sub.groupby(["Year", "Month"])["KMT"]
        .sum()
        .unstack("Month")
        .reindex(columns=range(1, 13))
    )
    piv.columns    = MONTH_ABBR
    piv.index      = piv.index.astype(int)
    piv.index.name = "Year"
    return piv



# =============================================================================
# DESTINATIONS & REGIONS BUILDERS
# =============================================================================

def assign_crop_year(df: pd.DataFrame, start_month: int) -> pd.DataFrame:
    """Adds CropYear column e.g. '12/13' based on Year, Month, start_month."""
    df = df.copy()
    cy_start = df["Year"].where(df["Month"] >= start_month, df["Year"] - 1)
    df["CropYear"] = cy_start.astype(str).str[-2:] + "/" + (cy_start + 1).astype(str).str[-2:]
    return df


def build_destinations(df: pd.DataFrame, trade_type: str,
                       crop_start: int, top_n: int) -> pd.DataFrame:
    """
    Returns a DataFrame:
      Rows    = crop years (sorted ascending)
      Columns = Top-N countries by total volume + "Rest of World" + "TOTAL"
      Values  = KMT (rounded to whole integers)
    Ranked by total volume across the full dataset.
    """
    sub = df[df["TradeType"] == trade_type].copy()
    if sub.empty:
        return pd.DataFrame()

    sub = assign_crop_year(sub, crop_start)

    # Rank countries by total volume across all years
    top_countries = (
        sub.groupby("Country")["KMT"].sum()
        .nlargest(top_n).index.tolist()
    )

    # Crop-year totals per country
    cy_country = (
        sub.groupby(["CropYear", "Country"])["KMT"]
        .sum()
        .reset_index()
    )

    # Pivot: rows=CropYear, cols=Country
    piv = cy_country.pivot(index="CropYear", columns="Country", values="KMT").fillna(0)

    # Sorted crop years
    all_cy = sorted(piv.index.tolist(),
                    key=lambda s: int(s.split("/")[0]))
    piv = piv.reindex(all_cy)

    # Top-N columns
    top_cols = [c for c in top_countries if c in piv.columns]
    rest_cols = [c for c in piv.columns if c not in top_cols]

    out = piv[top_cols].copy()
    out["Rest of World"] = piv[rest_cols].sum(axis=1)
    out["TOTAL"]         = piv.sum(axis=1)
    out.index.name       = "CropYear"

    # Transpose: countries as rows, crop years as columns
    # Note: TOTAL was already a column before transpose — it becomes a row here.
    # Do NOT add another out.sum() — that would double-count it.
    out = out.T
    out.index.name = "Country"

    return out.round(0).astype("Int64", errors="ignore")


def build_regions(df: pd.DataFrame, trade_type: str,
                  crop_start: int) -> pd.DataFrame:
    """
    Returns a DataFrame:
      Rows    = crop years (sorted ascending)
      Columns = regions in REGION_ORDER
      Values  = KMT (rounded to whole integers)
    """
    sub = df[df["TradeType"] == trade_type].copy()
    if sub.empty:
        return pd.DataFrame()

    sub = assign_crop_year(sub, crop_start)
    sub["Region"] = sub["Country"].map(REGION_MAP).fillna("Other")

    cy_region = (
        sub.groupby(["CropYear", "Region"])["KMT"]
        .sum()
        .reset_index()
    )

    piv = cy_region.pivot(index="CropYear", columns="Region", values="KMT").fillna(0)

    all_cy = sorted(piv.index.tolist(),
                    key=lambda s: int(s.split("/")[0]))
    piv = piv.reindex(all_cy)

    # Enforce region column order, only include regions present in data
    ordered_cols = [r for r in REGION_ORDER if r in piv.columns]
    piv = piv[ordered_cols]
    piv["TOTAL"] = piv.sum(axis=1)
    piv.index.name = "CropYear"

    # Transpose: regions as rows, crop years as columns
    piv = piv.T
    piv.index.name = "Region"

    return piv.round(0).astype("Int64", errors="ignore")

# =============================================================================
# EXCEL WRITER
# =============================================================================

# Colours
C = {
    "dark":      "1F4E79",
    "exp_hdr":   "1E5C1E",   # dark green for export headers
    "imp_hdr":   "7B3000",   # dark amber for import headers
    "exp_yr":    "C6EFCE",   # light green  — year cell
    "imp_yr":    "FFEB9C",   # light yellow — year cell
    "exp_data":  "EBF5EB",   # very light green — data rows
    "imp_data":  "FFFDE7",   # very light yellow
    "tot_fill":  "D9E1F2",   # blue tint — Annual column
    "alt":       "F2F2F2",   # alternating row
    "white":     "FFFFFF",
}

_THIN = Side(style="thin",   color="BBBBBB")
_MED  = Side(style="medium", color="888888")
BOX   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BOX_L = Border(left=_MED,  right=_THIN, top=_THIN, bottom=_THIN)  # left edge emphasis


def _hdr(ws, row, col, value, bg, fg="FFFFFF", align="center", size=10):
    c = ws.cell(row, col, value)
    c.font      = Font(bold=True, color=fg, name="Calibri", size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BOX
    return c


def _num(ws, row, col, value, bg):
    c = ws.cell(row, col)
    # Round to whole integer; leave blank if zero or NaN
    if pd.notna(value) and value != 0:
        c.value         = int(round(float(value)))
        c.number_format = "#,##0"
    else:
        c.value = None
    c.font      = Font(name="Calibri", size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right")
    c.border    = BOX
    return c


def write_pivot_sheet(wb: Workbook, sheet_name: str,
                      piv: pd.DataFrame, trade_type: str):
    if piv is None or piv.empty:
        return

    ws     = wb.create_sheet(sheet_name)
    bg_hdr = C["exp_hdr"]  if trade_type == "Export" else C["imp_hdr"]
    bg_yr  = C["exp_yr"]   if trade_type == "Export" else C["imp_yr"]
    bg_dat = C["exp_data"] if trade_type == "Export" else C["imp_data"]

    # ── Row 1: title spanning the full width
    total_cols = 1 + 12 + 1   # Year + 12 months + Annual
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=total_cols)
    tc = ws.cell(1, 1, f"Sum of KMT  —  {sheet_name}  |  Months (Period)")
    tc.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    tc.fill      = PatternFill("solid", fgColor=C["dark"])
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── Row 2: column headers
    _hdr(ws, 2, 1, "Years (Period)", bg_hdr, align="left")
    for ci, m in enumerate(MONTH_ABBR, 2):
        _hdr(ws, 2, ci, m, bg_hdr)
    _hdr(ws, 2, total_cols, "Annual", C["dark"])

    # ── Data rows
    for ri, (year, row) in enumerate(piv.iterrows(), 3):
        alt = bg_dat if ri % 2 != 0 else C["white"]

        # Year label cell
        yc = ws.cell(ri, 1, int(year))
        yc.font      = Font(bold=True, name="Calibri", size=10)
        yc.fill      = PatternFill("solid", fgColor=bg_yr)
        yc.alignment = Alignment(horizontal="center", vertical="center")
        yc.border    = BOX_L

        annual = 0
        for ci, m in enumerate(MONTH_ABBR, 2):
            v = row.get(m)
            _num(ws, ri, ci, v, alt)
            if pd.notna(v):
                annual += float(v)

        # Annual total
        ac = ws.cell(ri, total_cols)
        ac.value         = int(round(annual)) if annual else None
        ac.number_format = "#,##0"
        ac.font          = Font(bold=True, name="Calibri", size=10)
        ac.fill          = PatternFill("solid", fgColor=C["tot_fill"])
        ac.alignment     = Alignment(horizontal="right", vertical="center")
        ac.border        = BOX

    # ── Column widths
    ws.column_dimensions["A"].width = 16   # "Years (Period)"
    for ci in range(2, total_cols):
        ws.column_dimensions[get_column_letter(ci)].width = 8
    ws.column_dimensions[get_column_letter(total_cols)].width = 10

    ws.row_dimensions[2].height = 16
    ws.freeze_panes = "B3"


def write_raw_sheet(wb: Workbook, df: pd.DataFrame, commodity: str):
    """
    Flat table of every row for this commodity.
    Use the column autofilters to slice by Country, HS6, TradeType, Year, Month.
    This replaces the pivot slicer functionality.
    """
    ws   = wb.create_sheet("Raw Data")
    cols = ["Year", "Month", "TradeType", "HS6", "Country", "KMT", "ValueCAD"]

    # ── Header
    for ci, h in enumerate(cols, 1):
        _hdr(ws, 1, ci, h, C["dark"], size=10)

    # ── Data
    for ri, (_, row) in enumerate(df[cols].sort_values(
            ["TradeType", "Year", "Month", "Country"]).iterrows(), 2):
        alt = C["alt"] if ri % 2 == 0 else C["white"]
        for ci, col in enumerate(cols, 1):
            v    = row[col]
            cell = ws.cell(ri, ci)
            cell.fill   = PatternFill("solid", fgColor=alt)
            cell.border = BOX
            cell.font   = Font(name="Calibri", size=10)
            if col in ("Year", "Month"):
                cell.value      = int(v) if pd.notna(v) else None
                cell.alignment  = Alignment(horizontal="center")
            elif col == "HS6":
                cell.value      = int(v) if pd.notna(v) else None
                cell.alignment  = Alignment(horizontal="center")
            elif col == "KMT":
                cell.value      = round(float(v), 3) if pd.notna(v) and v else None
                cell.number_format = "#,##0.000"
                cell.alignment  = Alignment(horizontal="right")
            elif col == "ValueCAD":
                cell.value      = int(v) if pd.notna(v) and v else None
                cell.number_format = "#,##0"
                cell.alignment  = Alignment(horizontal="right")
            else:
                cell.value = str(v) if pd.notna(v) else None

    # ── Autofilter + freeze
    last_col = get_column_letter(len(cols))
    last_row = len(df) + 1
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes    = "A2"

    # ── Column widths
    widths = {
        "Year": 7, "Month": 7, "TradeType": 10,
        "HS6": 10, "Country": 28, "KMT": 12, "ValueCAD": 14
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 10)

    # ── Usage note in cell after the table
    note_row = last_row + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row,   end_column=5)
    nc = ws.cell(note_row, 1,
        "Filter by Country and/or HS6 to slice trade flows. "
        "KMT = thousands of metric tonnes.")
    nc.font      = Font(italic=True, color="888888", name="Calibri", size=9)
    nc.alignment = Alignment(horizontal="left")



def write_destinations_sheet(wb: Workbook, commodity: str,
                              dest_df: pd.DataFrame, reg_df: pd.DataFrame,
                              trade_type: str):
    """
    Writes two tables to the Destinations sheet.
    Layout (transposed): Rows = Country/Region, Columns = Crop Years
    Block 1 — Top-N countries by crop year
    Block 2 — Regional breakdown by crop year
    """
    if (dest_df is None or dest_df.empty) and (reg_df is None or reg_df.empty):
        return

    sheet_label = f"Dest ({trade_type[:3]})"
    ws   = wb.create_sheet(sheet_label)
    bg_h = C["exp_hdr"] if trade_type == "Export" else C["imp_hdr"]
    bg_f = C["exp_yr"]  if trade_type == "Export" else C["imp_yr"]
    bg_d = C["exp_data"] if trade_type == "Export" else C["imp_data"]

    def write_block(df, start_row, block_title, row_label):
        """
        df:  rows = Country or Region, cols = crop year labels + TOTAL
        row_label: header for the first column e.g. "Country" / "Region"
        """
        if df is None or df.empty:
            return start_row

        cy_cols   = [c for c in df.columns if c != "TOTAL"]
        has_total = "TOTAL" in df.columns
        total_cols = 1 + len(cy_cols) + (1 if has_total else 0)

        # ── Title bar
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=total_cols)
        tc = ws.cell(start_row, 1)
        tc.value     = block_title
        tc.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        tc.fill      = PatternFill("solid", fgColor=C["dark"])
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 20

        # ── Header row: row_label | 12/13 | 13/14 | ... | TOTAL
        _hdr(ws, start_row + 1, 1, row_label, bg_h, align="left")
        for ci, cy in enumerate(cy_cols, 2):
            _hdr(ws, start_row + 1, ci, str(cy), bg_h, align="center")
        if has_total:
            _hdr(ws, start_row + 1, total_cols, "TOTAL", C["dark"], align="center")

        # ── Data rows (one per country / region)
        for ri, (label, row) in enumerate(df.iterrows(), start_row + 2):
            is_total_row = str(label) == "TOTAL"
            alt = C["tot_fill"] if is_total_row else (bg_d if ri % 2 != 0 else C["white"])

            # Row label cell (Country or Region name)
            lc = ws.cell(ri, 1, str(label))
            lc.font      = Font(bold=is_total_row, name="Calibri", size=10)
            lc.fill      = PatternFill("solid", fgColor=bg_f if is_total_row else alt)
            lc.alignment = Alignment(horizontal="left")
            lc.border    = BOX

            # Crop year value cells
            for ci, cy in enumerate(cy_cols, 2):
                v    = row.get(cy)
                cell = ws.cell(ri, ci)
                try:
                    fv = int(v) if pd.notna(v) and float(v) != 0 else None
                except (TypeError, ValueError):
                    fv = None
                cell.value         = fv
                cell.number_format = "#,##0"
                cell.font          = Font(bold=is_total_row, name="Calibri", size=10)
                cell.fill          = PatternFill("solid", fgColor=alt)
                cell.alignment     = Alignment(horizontal="right")
                cell.border        = BOX

            # TOTAL column
            if has_total:
                v    = row.get("TOTAL")
                cell = ws.cell(ri, total_cols)
                try:
                    fv = int(v) if pd.notna(v) and float(v) != 0 else None
                except (TypeError, ValueError):
                    fv = None
                cell.value         = fv
                cell.number_format = "#,##0"
                cell.font          = Font(bold=True, name="Calibri", size=10)
                cell.fill          = PatternFill("solid", fgColor=C["tot_fill"])
                cell.alignment     = Alignment(horizontal="right")
                cell.border        = BOX

        # ── Column widths
        ws.column_dimensions["A"].width = 28   # country / region name
        for ci in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 9

        return start_row + 2 + len(df) + 2

    # ── Write both blocks
    next_row = write_block(
        dest_df, 1,
        f"Canada {trade_type}s — {commodity} (KMT)  |  Top Destinations by Crop Year",
        "Country"
    )
    write_block(
        reg_df, next_row,
        f"Canada {trade_type}s — {commodity} (KMT)  |  Regional Breakdown by Crop Year",
        "Region"
    )

    ws.freeze_panes = "B3"

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("  CIMT S&D Builder")
    print("=" * 60 + "\n")

    df_all = load_all_zips(CIMTD)
    today  = date.today().strftime("%Y%m%d")

    for commodity, cfg in COMMODITY_CONFIG.items():
        print(f"── {commodity}")
        df_c = df_all[df_all["HS6"].isin(cfg["hs6_codes"])].copy()

        if df_c.empty:
            print("   No data — check hs6_codes in config.\n")
            continue

        n_exp = (df_c["TradeType"] == "Export").sum()
        n_imp = (df_c["TradeType"] == "Import").sum()
        print(f"   {len(df_c):,} rows  ({n_exp:,} export / {n_imp:,} import)")

        crop_start = cfg["crop_year_start_month"]
        top_n      = cfg.get("top_destinations", 10)

        piv_exp  = build_pivot(df_c, "Export")
        piv_imp  = build_pivot(df_c, "Import")
        dest_exp = build_destinations(df_c, "Export", crop_start, top_n)
        dest_imp = build_destinations(df_c, "Import", crop_start, top_n)
        reg_exp  = build_regions(df_c, "Export", crop_start)
        reg_imp  = build_regions(df_c, "Import", crop_start)

        wb = Workbook()
        wb.remove(wb.active)
        write_pivot_sheet(wb, "Exports", piv_exp, "Export")
        write_pivot_sheet(wb, "Imports", piv_imp, "Import")
        write_destinations_sheet(wb, commodity, dest_exp, reg_exp, "Export")
        write_destinations_sheet(wb, commodity, dest_imp, reg_imp, "Import")
        write_raw_sheet(wb, df_c, commodity)

        safe_name = commodity.replace("/", "-").replace(" ", "_")
        fname     = CIMTD / f"CIMT_{safe_name}_{today}.xlsx"
        wb.save(fname)
        print(f"   Saved -> {fname.name}\n")

    print("Done.")


if __name__ == "__main__":
    main()
