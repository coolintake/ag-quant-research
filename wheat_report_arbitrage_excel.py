"""
report_arbitrage_excel.py — Global Wheat Arbitrage Matrix as a styled Excel file.

Reads from global_wheat_summary.db and writes a presentation-ready .xlsx.
The file is designed to be pasted directly into PowerPoint as a table or image.

Usage:
    python report_arbitrage_excel.py
    python report_arbitrage_excel.py --db PATH/TO/global_wheat_summary.db
    python report_arbitrage_excel.py --out my_matrix.xlsx

Output: arbitrage_matrix_YYYYMMDD.xlsx  (or --out filename)
"""

import sqlite3
import sys
import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import DB_SUMMARY
from mappings import (
    MATRIX_COMMODITIES, MATRIX_BASELINES, SPOT_WINDOW,
)
from report_utils import get_comparison_window, derive_display_windows

# ── Colour palette ─────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"   # header background
C_MID_BLUE    = "2E5FA3"   # tier label background
C_LIGHT_BLUE  = "D6E4F0"   # column header background
C_BASELINE    = "FFF2CC"   # Canada baseline row highlight
C_CHEAPER     = "E2EFDA"   # green  — competitor cheaper than Canada (threat)
C_DEARER      = "FCE4D6"   # orange — competitor more expensive (Canada competitive)
C_NEUTRAL     = "FFFFFF"   # no spread (low-pro tier)
C_BORDER      = "B8CCE4"
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"
C_DARK_GREEN  = "375623"
C_DARK_RED    = "843C0C"
C_DARK_GOLD   = "7F6000"

THIN  = Side(style="thin",   color=C_BORDER)
THICK = Side(style="medium", color=C_MID_BLUE)
BORDER_THIN  = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
BORDER_THICK_BOT = Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

FONT_TITLE    = Font(name="Arial", bold=True,  size=13, color=C_WHITE)
FONT_TIER     = Font(name="Arial", bold=True,  size=10, color=C_WHITE)
FONT_COL_HDR  = Font(name="Arial", bold=True,  size=9,  color=C_DARK_BLUE)
FONT_BASELINE = Font(name="Arial", bold=True,  size=9,  color=C_DARK_GOLD)
FONT_BODY     = Font(name="Arial", bold=False, size=9,  color=C_BLACK)
FONT_SPREAD_POS = Font(name="Arial", bold=True, size=9, color=C_DARK_RED)
FONT_SPREAD_NEG = Font(name="Arial", bold=True, size=9, color=C_DARK_GREEN)
FONT_META     = Font(name="Arial", italic=True, size=8, color="595959")
FONT_FOOTNOTE = Font(name="Arial", italic=True, size=8, color="595959")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Data loading (same logic as report_arbitrage.py) ──────────────────────────

def load_prices(db_path: Path, display_windows: list[str]) -> dict:
    active = {k: v for k, v in MATRIX_COMMODITIES.items()
              if not v.get("suppress", False)}
    con  = sqlite3.connect(db_path)
    rows = con.execute(
        "SELECT Origin, Commodity, Delivery_Window, Price_USD "
        "FROM wheat_summary WHERE Price_USD IS NOT NULL"
    ).fetchall()
    con.close()

    prices = {}
    for origin, commodity, window, price in rows:
        if commodity not in active:
            continue
        key = (origin, commodity)
        if key not in prices:
            prices[key] = {}
        prompt_col = display_windows[0] if display_windows else "Jul-26"
        col = prompt_col if window == SPOT_WINDOW else window
        if window == SPOT_WINDOW:
            prices[key]["_spot"] = True
        prices[key][col] = price
    return prices


def load_metadata(db_path: Path) -> dict:
    try:
        con = sqlite3.connect(db_path)
        row = con.execute(
            "SELECT usdcad, prompt_window, report_date "
            "FROM run_metadata ORDER BY run_at DESC LIMIT 1"
        ).fetchone()
        con.close()
        if row:
            return {"usdcad": row[0], "prompt_window": row[1], "report_date": row[2]}
    except Exception:
        pass
    return {}


def get_report_date(meta: dict) -> str:
    rd = meta.get("report_date")
    if not rd:
        return datetime.now().strftime("%B %d, %Y")
    try:
        return datetime.strptime(rd, "%Y-%m-%d").strftime("%B %d, %Y")
    except Exception:
        return rd


def _find_prices(prices, commodity, origin_short):
    for (origin, comm), wp in prices.items():
        if comm == commodity and origin_short.lower() in origin.lower():
            return wp
    for (origin, comm), wp in prices.items():
        if comm == commodity:
            return wp
    return {}


# ── Excel writer ───────────────────────────────────────────────────────────────

def build_excel(db_path: Path, out_path: Path):
    comparison_window = get_comparison_window(db_path)
    windows           = derive_display_windows(db_path, comparison_window)
    spread_window     = comparison_window
    prices            = load_prices(db_path, windows)
    meta   = load_metadata(db_path)
    report_date = get_report_date(meta)
    usdcad      = meta.get("usdcad")

    wb = Workbook()
    ws = wb.active
    ws.title = "Arbitrage Matrix"

    # Column layout: A=Origin B=Grade C=Jun D=Jul E=Aug F=vs Jul
    # Widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    row = 1

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value     = "GLOBAL WHEAT ARBITRAGE MATRIX"
    c.font      = FONT_TITLE
    c.fill      = _fill(C_DARK_BLUE)
    c.alignment = _center()
    ws.row_dimensions[row].height = 22
    row += 1

    # Meta rows
    meta_lines = [
        f"Report Date: {report_date}    |    Units: USD/MT FOB    |    USD/CAD: {usdcad:.4f}" if usdcad else f"Report Date: {report_date}    |    Units: USD/MT FOB",
        "Spread = Competitor Jul-26 price minus Canadian baseline Jul-26.  Negative (green) = cheaper than Canada.  Positive (orange) = more expensive.",
    ]
    for line in meta_lines:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value     = line
        c.font      = FONT_META
        c.fill      = _fill("EBF3FB")
        c.alignment = _left(wrap=True)
        ws.row_dimensions[row].height = 14
        row += 1

    row += 1  # blank spacer

    # ── Render each tier ──────────────────────────────────────────────────────
    for tier in ["High", "Mid", "Low"]:
        baseline_cfg = MATRIX_BASELINES[tier]
        has_baseline = baseline_cfg is not None
        active = {k: v for k, v in MATRIX_COMMODITIES.items()
                  if v.get("tier") == tier and not v.get("suppress", False)}

        # Tier label row
        tier_labels = {"High": "HIGH PROTEIN", "Mid": "MID PROTEIN",
                       "Low": "LOW PROTEIN  (reference only — no Canadian baseline)"}
        if has_baseline:
            b_cfg  = MATRIX_COMMODITIES.get(baseline_cfg["commodity"], {})
            b_note = f"  |  Baseline: {b_cfg.get('origin_short','')} — {baseline_cfg['commodity']}"
        else:
            b_note = ""

        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value     = tier_labels[tier] + b_note
        c.font      = FONT_TIER
        c.fill      = _fill(C_MID_BLUE)
        c.alignment = _left()
        ws.row_dimensions[row].height = 16
        row += 1

        # Column header row
        headers = ["Origin", "Grade"] + windows + ([f"vs {spread_window}"] if has_baseline else [])
        for col_idx, hdr in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx)
            c.value     = hdr
            c.font      = FONT_COL_HDR
            c.fill      = _fill(C_LIGHT_BLUE)
            c.alignment = _center()
            c.border    = BORDER_THIN
        ws.row_dimensions[row].height = 14
        row += 1

        # Baseline Jul-26 price for spread
        baseline_jul = None
        if has_baseline:
            wp = _find_prices(prices, baseline_cfg["commodity"],
                              MATRIX_COMMODITIES[baseline_cfg["commodity"]]["origin_short"])
            baseline_jul = wp.get("Jul-26")

        # Collect + sort rows
        data_rows    = []
        baseline_row = None

        for commodity, cfg in active.items():
            wp      = _find_prices(prices, commodity, cfg["origin_short"])
            if not wp:
                continue
            is_spot = wp.get("_spot", False)
            win_prices   = [wp.get(w) for w in windows]
            spread_price = wp.get(spread_window)
            spread = round(spread_price - baseline_jul, 2) if (
                has_baseline and baseline_jul is not None and spread_price is not None
            ) else None

            entry = (cfg["origin_short"], cfg["display_name"],
                     win_prices, spread, is_spot, commodity)
            if has_baseline and commodity == baseline_cfg["commodity"]:
                baseline_row = entry
            else:
                data_rows.append(entry)

        data_rows.sort(key=lambda r: (r[3] is None, r[3] or 9999))
        if baseline_row:
            data_rows.append(baseline_row)

        # Write data rows
        for entry in data_rows:
            origin_s, grade_s, win_prices, spread, is_spot, commodity = entry
            is_baseline = has_baseline and commodity == baseline_cfg["commodity"]

            # Row background
            if is_baseline:
                row_fill = _fill(C_BASELINE)
            elif spread is None:
                row_fill = _fill(C_NEUTRAL)
            elif spread < 0:
                row_fill = _fill(C_CHEAPER)    # green — cheaper than Canada
            else:
                row_fill = _fill(C_DEARER)     # orange — more expensive

            # Origin
            c = ws.cell(row=row, column=1, value=origin_s)
            c.font      = FONT_BASELINE if is_baseline else FONT_BODY
            c.fill      = row_fill
            c.alignment = _left()
            c.border    = BORDER_THIN

            # Grade
            c = ws.cell(row=row, column=2, value=grade_s)
            c.font      = FONT_BASELINE if is_baseline else FONT_BODY
            c.fill      = row_fill
            c.alignment = _left()
            c.border    = BORDER_THIN

            # Price columns
            for col_idx, val in enumerate(win_prices, 3):
                suffix = "*" if (col_idx == 3 and is_spot) else ""
                display = f"${val:.2f}{suffix}" if val is not None else "—"
                c = ws.cell(row=row, column=col_idx, value=display)
                c.font      = FONT_BASELINE if is_baseline else FONT_BODY
                c.fill      = row_fill
                c.alignment = _center()
                c.border    = BORDER_THIN

            # Spread column
            if has_baseline:
                if is_baseline:
                    spread_val = "BASELINE"
                    sfont      = FONT_BASELINE
                elif spread is None:
                    spread_val = "—"
                    sfont      = FONT_BODY
                elif spread < 0:
                    spread_val = f"{spread:+.2f}"
                    sfont      = Font(name="Arial", bold=True, size=9,
                                      color=C_DARK_GREEN)
                else:
                    spread_val = f"+{spread:.2f}"
                    sfont      = Font(name="Arial", bold=True, size=9,
                                      color=C_DARK_RED)

                c = ws.cell(row=row, column=6, value=spread_val)
                c.font      = sfont
                c.fill      = row_fill
                c.alignment = _center()
                c.border    = BORDER_THIN

            ws.row_dimensions[row].height = 14
            row += 1

        row += 1  # spacer between tiers

    # ── Footnotes ─────────────────────────────────────────────────────────────
    notes = [
        "* Spot indication — not a named forward month.",
        "Canadian grades: 13.5% moisture basis.  US grades: 12% moisture basis.  International (Hammer): typically DMB.",
    ]
    for note in notes:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value     = note
        c.font      = FONT_FOOTNOTE
        c.alignment = _left()
        ws.row_dimensions[row].height = 12
        row += 1

    # ── Freeze panes below title/meta ─────────────────────────────────────────
    ws.freeze_panes = "A5"

    wb.save(out_path)
    print(f"  Saved → {out_path}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export wheat arbitrage matrix as a styled Excel file"
    )
    parser.add_argument("--db",  type=str, default=None)
    parser.add_argument("--out", type=str, default=None)
    args = parser.parse_args()

    db_path  = Path(args.db)  if args.db  else DB_SUMMARY
    ts       = datetime.now().strftime("%Y%m%d")
    out_path = Path(args.out) if args.out else Path(__file__).parent / f"arbitrage_matrix_{ts}.xlsx"

    if not db_path.exists():
        print(f"Error: database not found: {db_path}")
        print("Run python run_pipeline.py first.")
        sys.exit(1)

    build_excel(db_path, out_path)


if __name__ == "__main__":
    main()
