import time
import threading
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
from scipy.stats import norm
from scipy.optimize import brentq
from scipy.interpolate import griddata
import os
import openpyxl
from openpyxl import load_workbook

from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.contract import Contract

print("✓ Libraries imported successfully")

# ===== FINANCIAL MATH (BLACK-76) =====
class VolMath:
    @staticmethod
    def black_76_price(F, K, T, r, sigma, option_type="C"):
        """
        Black-76 for Futures Options.
        F: Futures Price
        K: Strike
        T: Time to Expiry (years)
        r: Risk-free rate
        sigma: Volatility
        """
        if T <= 0:
            return max(0, F - K) if option_type == "C" else max(0, K - F)
            
        d1 = (np.log(F / K) + (sigma**2 / 2) * T) / (sigma * np.sqrt(T))
        d2 = d1 - sigma * np.sqrt(T)
        
        disc = np.exp(-r * T)
        
        if option_type == "C":
            return disc * (F * norm.cdf(d1) - K * norm.cdf(d2))
        else:
            return disc * (K * norm.cdf(-d2) - F * norm.cdf(-d1))

    @staticmethod
    def solve_iv(market_price, F, K, T, r, option_type="C"):
        """
        Solve for Implied Volatility using Black-76
        """
        # Intrinsic boundary check
        intrinsic = max(0, F - K) if option_type == "C" else max(0, K - F)
        if market_price <= intrinsic:
            return np.nan # Arbitrage violation or deep ITM illiquidity
        if T <= 0:
            return np.nan

        def objective(sigma):
            return VolMath.black_76_price(F, K, T, r, sigma, option_type) - market_price

        try:
            # Brent's method is robust for IV solving
            return brentq(objective, 1e-4, 4.0, xtol=1e-5)
        except Exception:
            return np.nan

# ===== ROBUST IBKR CONNECTOR =====
class AsyncCornApp(EClient, EWrapper):
    def __init__(self):
        EClient.__init__(self, self)
        EWrapper.__init__(self)
        
        # State
        self.nextOrderId = None
        self.connected_event = threading.Event()
        
        # Data
        self.underlying_price = None
        self.underlying_event = threading.Event()
        
        # Contracts
        self.found_contracts = []
        self.contract_details_list = []  # Store full contractDetails objects
        self.contract_details_event = threading.Event()
        self.future_contract_event = threading.Event()
        self.future_contract_symbol = None
        
        # Batch Data
        self.market_data = {}  # reqId -> dict
        
    def nextValidId(self, orderId):
        super().nextValidId(orderId)
        self.nextOrderId = orderId
        self.connected_event.set()
        print(f"✓ Connected (NextOID: {orderId})")

    def error(self, reqId, errorCode, errorString, adv=None):
        if errorCode in [2104, 2106, 2158, 2157]:
            pass # Benign
        else:
            # print(f"Error {errorCode}: {errorString}") # Quiet errors for cleaner output
            pass

    def contractDetails(self, reqId, contractDetails):
        self.found_contracts.append(contractDetails.contract)
        self.contract_details_list.append(contractDetails)  # Store full details
    
    def contractDetailsEnd(self, reqId):
        # Always set the event for contract requests
        self.contract_details_event.set()

    def tickPrice(self, reqId, tickType, price, attrib):
        if price <= 0: return

        # 1001 = Underlying Request ID
        if reqId == 1001 and tickType in [1, 2, 4]:
            self.underlying_price = price
            self.underlying_event.set()
            
        # Option Prices
        elif reqId in self.market_data:
            if tickType == 1: self.market_data[reqId]['bid'] = price
            elif tickType == 2: self.market_data[reqId]['ask'] = price
            elif tickType == 4: self.market_data[reqId]['last'] = price

    def tickOptionComputation(self, reqId, tickType, tickAttrib, impliedVol, delta, optPrice, pvDividend, gamma, vega, theta, undPrice):
        # 13 = Model Option
        if reqId in self.market_data:
            if impliedVol and impliedVol > 0:
                self.market_data[reqId]['ib_iv'] = impliedVol
    
    def contractDetailsEnd(self, reqId):
        # Handle both future and option contract requests
        if reqId == 9999:  # Future contract verification
            self.future_contract_event.set()
        else:
            super().contractDetailsEnd(reqId)


class VolatilityManager:
    def __init__(self):
        self.app = AsyncCornApp()
        self.risk_free_rate = 0.045 # 4.5% assumed
    
    def connect(self):
        print("Connecting to TWS...")
        self.app.connect("127.0.0.1", 7496, 422)
        threading.Thread(target=self.app.run, daemon=True).start()
        if not self.app.connected_event.wait(10):
            raise Exception("Connection Timeout")
        self.app.reqMarketDataType(2) # Frozen

    def get_future_price_by_conid(self, conid):
        """
        Get futures price using the contract ID from the option's underlying.
        This ensures we get the EXACT futures contract the options are written on.
        """
        c = Contract()
        c.conId = conid
        c.exchange = "CBOT"
        
        print(f"Requesting futures price for ConID {conid}...")
        self.app.underlying_event.clear()
        self.app.reqMktData(1001, c, "", False, False, [])
        
        if self.app.underlying_event.wait(10):
            print(f"✓ Futures Price: {self.app.underlying_price}")
            self.app.cancelMktData(1001)
            return self.app.underlying_price
        else:
            print(f"✗ Failed to get futures price")
            return None

    def get_chain(self, expiry):
        c = Contract()
        c.symbol = "ZC"
        c.secType = "FOP"
        c.exchange = "CBOT"
        c.currency = "USD"
        c.lastTradeDateOrContractMonth = expiry
        
        self.app.found_contracts = []
        self.app.contract_details_list = []
        self.app.contract_details_event.clear()
        
        # print(f"Scanning {expiry}...")
        self.app.reqContractDetails(self.app.nextOrderId, c)
        self.app.nextOrderId += 1
        
        if self.app.contract_details_event.wait(30):
            calls = [co for co in self.app.found_contracts if co.right == 'C']
            details_calls = [cd for cd in self.app.contract_details_list if cd.contract.right == 'C']
            
            # Filter to select longest duration option expiry when multiple exist
            if calls:
                # Extract unique expiration dates
                expiration_dates = {co.lastTradeDateOrContractMonth for co in calls}
                
                # Select the longest duration (latest expiry)
                latest_expiry = max(expiration_dates)
                
                if len(expiration_dates) > 1:
                    print(f"  Found {len(expiration_dates)} option expiries: {sorted(expiration_dates)}")
                    print(f"  Selected longest duration: {latest_expiry}")
                
                # Filter to only contracts with the latest expiry
                calls = [co for co in calls if co.lastTradeDateOrContractMonth == latest_expiry]
                details_calls = [cd for cd in details_calls if cd.contract.lastTradeDateOrContractMonth == latest_expiry]
                
                # Extract underlying futures contract from first option's details
                underlying_conid = None
                if details_calls:
                    # The contractDetails has underlyingConId which identifies the futures contract
                    underlying_conid = details_calls[0].underlyingConId
            
            # Dedup by strike
            d = {x.strike: x for x in calls}
            return sorted(list(d.values()), key=lambda x: x.strike), latest_expiry, underlying_conid
        return [], None, None

    def get_data_for_chain(self, contracts, underlying_price, expiry_date_str):
        if not contracts: return []
        
        current_req_start = self.app.nextOrderId
        
        # Batch Request
        for i, c in enumerate(contracts):
            rid = current_req_start + i
            self.app.market_data[rid] = {
                'bid': None, 'ask': None, 'last': None, 'ib_iv': None,
                'strike': c.strike, 'contract': c, 'expiry': expiry_date_str,
                'underlying_price': underlying_price # Store relevant sport for this option
            }
            self.app.reqMktData(rid, c, "100,101,106", False, False, [])
            
        self.app.nextOrderId += len(contracts)
        
        # Wait for data aggregation
        time.sleep(4) 
        
        # Calculate & Cleanup
        results = []
        now = datetime.now()
        
        try:
            exp_dt = datetime.strptime(expiry_date_str, "%Y%m%d")
        except:
            exp_dt = datetime.strptime(expiry_date_str + "15", "%Y%m%d")
            
        T = (exp_dt - now).days / 365.0
        days_to_exp = (exp_dt - now).days
        
        for rid, item in self.app.market_data.items():
            if item['contract'] not in contracts: continue 
            
            self.app.cancelMktData(rid)
            
            bid = item['bid']
            ask = item['ask']
            last = item['last']
            
            mid_price = None
            if bid and ask and bid > 0 and ask > 0:
                mid_price = (bid + ask) / 2
            elif last and last > 0:
                mid_price = last
            
            # Calculate IV using SPECIFIC underlying
            calc_iv = np.nan
            if mid_price and underlying_price:
                calc_iv = VolMath.solve_iv(mid_price, underlying_price, item['strike'], T, self.risk_free_rate)
            
            final_iv = calc_iv if not np.isnan(calc_iv) else (item['ib_iv'] if item['ib_iv'] else np.nan)
            
            if not np.isnan(final_iv):
                results.append({
                    'strike': item['strike'],
                    'expiry': expiry_date_str, 
                    'days_to_expiry': days_to_exp,
                    'T': T,
                    'bid': bid, 'ask': ask, 'mid': mid_price,
                    'calc_iv': calc_iv,
                    'ib_iv': item['ib_iv'],
                    'iv': final_iv,
                    'underlying_price': underlying_price 
                })
        
        return results

    def disconnect(self):
        self.app.disconnect()


# ===== VALIDATION =====
def validate_surface(df):
    """
    Check for basic arbitrage violations.
    """
    print("\n--- Arbitrage Validation ---")
    
    # 1. Butterfly Arbitrage (Convexity check)
    butterfly_violations = 0
    total_checks_bf = 0
    
    for exp, group in df.groupby('expiry'):
        sorted_g = group.sort_values('strike')
        prices = sorted_g['mid'].values
        strikes = sorted_g['strike'].values
        
        for i in range(1, len(prices)-1):
            k1, k2, k3 = strikes[i-1], strikes[i], strikes[i+1]
            p1, p2, p3 = prices[i-1], prices[i], prices[i+1]
            if abs((k2-k1) - (k3-k2)) < 0.01:
                total_checks_bf += 1
                convexity = p1 - 2*p2 + p3
                if convexity < -0.1: 
                    butterfly_violations += 1
    
    print(f"Butterfly (Convexity) Violations: {butterfly_violations} / {total_checks_bf} checks")

    # 2. Calendar Arbitrage
    calendar_violations = 0
    total_checks_cal = 0
    
    df['total_var'] = (df['iv'] ** 2) * df['T']
    pivot = df.pivot_table(index='strike', columns='expiry', values='total_var')
    sorted_cols = sorted(pivot.columns)
    
    for i in range(len(sorted_cols)-1):
        t1 = pivot[sorted_cols[i]]
        t2 = pivot[sorted_cols[i+1]]
        diff = t2 - t1
        violations = (diff < -0.01).sum() 
        calendar_violations += violations
        total_checks_cal += diff.count()
        
    print(f"Calendar (Time) Violations:    {calendar_violations} / {total_checks_cal} checks")
    print("----------------------------")

# ===== VISUALIZATION =====
def visualize_surface(df):
    # 1. FILTERING (crucial for smooth surface)
    print(f"\nApplying filters...")
    
    # Filter A: Moneyness (row specific)
    df['moneyness'] = df['strike'] / df['underlying_price']
    mask_money = (df['moneyness'] >= 0.5) & (df['moneyness'] <= 1.6)
    
    # Filter B: IV Outliers 
    mask_iv = (df['iv'] < 2.5) & (df['iv'] > 0.01)
    
    clean = df[mask_money & mask_iv].copy()
    
    dropped = len(df) - len(clean)
    print(f"Dropped {dropped} data points. Remaining: {len(clean)}")

    if len(clean) < 10:
        print("Not enough data for surface after filtering.")
        return
        
    # --- STATISTICS ---
    print("\n--- Surface Statistics ---")
    print(clean['iv'].describe())
    
    # --- VALIDATION ---
    validate_surface(clean)

    # Create REGULAR GRID for surface
    # X = Strike, Y = Days
    unique_exp = clean['days_to_expiry'].nunique()
    if unique_exp < 2:
        print(f"\nWarning: Only {unique_exp} expiry found. Cannot interpolate 3D surface (need at least 2).")
        print("Showing 2D views only.")
    else:
        try:
            xi = np.linspace(clean['strike'].min(), clean['strike'].max(), 60)
            yi = np.linspace(clean['days_to_expiry'].min(), clean['days_to_expiry'].max(), 60)
            X, Y = np.meshgrid(xi, yi)
            
            Z = griddata(
                (clean['strike'], clean['days_to_expiry']), 
                clean['iv'], 
                (X, Y), 
                method='cubic'
            )
            
            # 1. 3D Surface
            fig = go.Figure(data=[go.Surface(x=X, y=Y, z=Z, colorscale='Viridis')])
            fig.update_layout(
                title=f'Corn (ZC) Volatility Surface<br>Black-76 (Term Structure Corrected)',
                scene=dict(
                    xaxis_title='Strike Price',
                    yaxis_title='Days to Expiry',
                    zaxis_title='Implied Volatility'
                ),
                width=1000, height=800
            )
            fig.show()
        except Exception as e:
            print(f"Interpolation failed: {e}")

    # 2. 2D Heatmap & Smiles
    fig2 = make_subplots(
        rows=2, cols=1, 
        specs=[[{"type": "xy"}], [{"type": "xy"}]],
        subplot_titles=("Volatility Heatmap", "Volatility Smiles")
    )
    
    piv = clean.pivot_table(index='days_to_expiry', columns='strike', values='iv')
    
    fig2.add_trace(go.Heatmap(
        z=clean['iv'],
        x=clean['strike'],
        y=clean['days_to_expiry'],
        colorscale='Viridis',
        colorbar=dict(title='IV')
    ), row=1, col=1)
    
    for exp, group in clean.groupby('expiry'):
        group = group.sort_values('strike')
        fig2.add_trace(go.Scatter(
            x=group['strike'], y=group['iv'],
            mode='lines+markers', name=str(exp)
        ), row=2, col=1)

    fig2.update_layout(height=900, title_text="Volatility Analysis Views")
    fig2.show()


# ===== MAIN =====
def main():
    vm = VolatilityManager()
    try:
        vm.connect()
        
        # TARGET EXPIRATIONS
        expiries = [
            '202603', '202605', '202607', 
            '202609', '202612'
        ]
        
        all_data = []
        
        for exp in expiries:
            print(f"\n--- Processing Futures Month: {exp} ---")
            
            # Get contracts (will auto-select longest duration and extract underlying ConID)
            contracts, actual_expiry, underlying_conid = vm.get_chain(exp)
            
            if not contracts:
                print(f"No contracts found for {exp}")
                continue
            
            # Log the actual option expiry that was selected
            print(f"✓ Selected option expiry: {actual_expiry}")
            print(f"  Found {len(contracts)} strikes.")
            
            # Get Future Price using the EXACT underlying contract from options
            if underlying_conid:
                print(f"  Underlying futures ConID: {underlying_conid}")
                fut_price = vm.get_future_price_by_conid(underlying_conid)
            else:
                print(f"  Warning: No underlying ConID found, cannot get futures price")
                fut_price = None
                
            if not fut_price or fut_price <= 0:
                print(f"Skipping {exp} (No Underlying Price)")
                continue
            
            data = vm.get_data_for_chain(contracts, fut_price, actual_expiry)
            print(f"Collected {len(data)} valid IV points.")
            all_data.extend(data)
            
        print(f"\nTotal Data Points: {len(all_data)}")
        
        df = pd.DataFrame(all_data)
        if not df.empty:
            # Save to Excel workbook with date-stamped worksheet
            excel_path = r"C:\Users\ahmed\OneDrive\Desktop\Python\Volatility_Surface_CZ\corn_options_surface_historical.xlsx"
            worksheet_name = f"ZC-{datetime.now().strftime('%d%m%y')}"
            
            # Load or create workbook
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                print(f"  Loading existing workbook")
                
                # Remove today's worksheet if it already exists
                if worksheet_name in wb.sheetnames:
                    wb.remove(wb[worksheet_name])
                    print(f"  Overwriting worksheet: {worksheet_name}")
            else:
                wb = openpyxl.Workbook()
                print(f"  Creating new workbook")
            
            # Create new worksheet and add data
            ws = wb.create_sheet(title=worksheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save workbook
            wb.save(excel_path)
            
            print(f"\n✓ Data saved to: {excel_path}")
            print(f"  Worksheet: {worksheet_name}")
            print(f"  Total worksheets: {len(wb.sheetnames)}")

            print("\nSample Data:")
            print(df[['expiry', 'strike', 'mid', 'calc_iv', 'ib_iv', 'iv', 'underlying_price']].head())
            visualize_surface(df)
            
    except Exception as e:
        print(f"Error: {e}")
    finally:
        vm.disconnect()

if __name__ == "__main__":
    main()
